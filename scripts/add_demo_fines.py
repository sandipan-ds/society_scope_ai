"""Add demo fines + payments for flat 104 so the agent has non-trivial
private data to answer with. Safe to inspect; re-run is idempotent (sets
absolute cell values, not increments).

Demo scenario for flat 104 (Shreya Sisodia):
  - Parking fine Rs 500 in Mar-26
  - Noise fine Rs 300 in Apr-26
  - Payments: Jan-26 = 3500 (paid), Feb-26 = 3500 (paid), Mar-26 = 2000 (partial)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(r"C:\Users\sandi\Desktop\ML Working Folder\society_scope_ai\data\members_data\Housing_Society_Charges_and_Fines_Template_108_Residents.xlsx")

FLAT = "104"

# sheet -> {month-column-header: value}
EDITS = {
    "Parking Violation Fines": {"Mar-26": 500},
    "Noise Violation Fines": {"Apr-26": 300},
    "Payments": {"Jan-26": 3500, "Feb-26": 3500, "Mar-26": 2000},
}


def main() -> None:
    wb = load_workbook(XLSX)
    for sheet, changes in EDITS.items():
        ws = wb[sheet]
        headers = {cell.value: cell.column for cell in ws[1]}
        flat_col = headers["Flat No."]
        target_row = None
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=flat_col).value).strip() == FLAT:
                target_row = row
                break
        if target_row is None:
            raise RuntimeError(f"Flat {FLAT} not found in {sheet}")
        for month, value in changes.items():
            ws.cell(row=target_row, column=headers[month]).value = value
        print(f"{sheet}: flat {FLAT} row {target_row} -> {changes}")
    wb.save(XLSX)
    print("Saved", XLSX)


if __name__ == "__main__":
    main()
