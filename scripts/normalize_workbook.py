"""Normalize the society workbook to match the documented model.

ONE-TIME SETUP SCRIPT — do NOT re-run after data entry has started.
It rebuilds the Payments sheet as all zeros and resets fine-sheet layout,
which would destroy any payment data already entered.

- Renames Maintenance month columns to Jan-26..Dec-26 (consistent with fine sheets)
- Adds Excel formulas so each Maintenance month = 3500 + sum of all fine sheets
- Adds a Payments sheet (zero-filled, same grid) for tracking paid amounts
- Removes blank rows from fine sheets
- Ensures all financial sheets cover the same 108 flats in Residents order
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd

XLSX = Path(r"C:\Users\sandi\Desktop\ML Working Folder\society_scope_ai\data\members_data\Housing_Society_Charges_and_Fines_Template_108_Residents.xlsx")

FINE_SHEETS = [
    "Parking Violation Fines",
    "Waste Management Fines",
    "Pet Policy Fines",
    "Noise Violation Fines",
    "Property Damage Fines",
    "Miscellaneous Fines",
]

MONTH_COLS = ["Jan-26", "Feb-26", "Mar-26", "Apr-26", "May-26", "Jun-26",
              "Jul-26", "Aug-26", "Sep-26", "Oct-26", "Nov-26", "Dec-26"]


def _month_letter(col_index_0: int) -> str:
    """Column I=8->'I', T=19->'T'."""
    return chr(ord("A") + col_index_0)


def main() -> None:
    sheets = pd.read_excel(XLSX, sheet_name=None, engine="openpyxl")

    residents = sheets["Residents"].copy()
    residents["Flat No."] = residents["Flat No."].astype(int)
    flat_list = residents["Flat No."].tolist()

    # Normalize Maintenance Charges
    mc = sheets["Maintenance Charges"].copy()
    mc["Flat No."] = mc["Flat No."].astype("Int64")
    # Rename inconsistent month columns to Jan-26..Dec-26
    old_month_cols = [c for c in mc.columns if c not in residents.columns]
    rename_map = dict(zip(old_month_cols, MONTH_COLS))
    mc = mc.rename(columns=rename_map)

    # Reindex to all 108 flats, filling missing with Residents info + 3500
    mc = mc.set_index("Flat No.").reindex(flat_list).reset_index()
    for col in ["Owner Name", "Resident Name", "Occupancy Type", "Owner Email",
                "Owner Mobile", "Resident Email", "Resident Mobile"]:
        if col in residents.columns:
            lookup = residents.set_index("Flat No.")[col]
            mc[col] = mc["Flat No."].map(lookup)

    # Add formulas: =3500 + sum of fine sheet cells
    # Month columns are I..T (0-indexed 8..19), rows 2..109 in Excel
    for month_idx, month in enumerate(MONTH_COLS):
        col_letter = _month_letter(8 + month_idx)
        refs = [f"'{name}'!{col_letter}{{row}}" for name in FINE_SHEETS]
        formulas = [f"=3500+{'+'.join(refs)}" for _ in range(len(mc))]
        # Store formulas as strings; pandas ExcelWriter can't write formulas,
        # so we'll overwrite with openpyxl below.
        mc[month] = formulas

    # Build Payments sheet (zero-filled, same identity columns + months)
    payments = residents.copy()
    for m in MONTH_COLS:
        payments[m] = 0

    # Clean fine sheets: drop fully empty rows, reindex to flat list
    for name in FINE_SHEETS:
        df = sheets[name].copy()
        df["Flat No."] = df["Flat No."].astype("Int64")
        df = df.dropna(subset=["Flat No."])
        df = df.set_index("Flat No.").reindex(flat_list).reset_index()
        # Fill identity columns from Residents
        for col in ["Owner Name", "Resident Name", "Occupancy Type", "Owner Email",
                    "Owner Mobile", "Resident Email", "Resident Mobile"]:
            if col in residents.columns:
                lookup = residents.set_index("Flat No.")[col]
                df[col] = df["Flat No."].map(lookup)
        # Fill any missing fine values with 0
        for m in MONTH_COLS:
            if m in df.columns:
                df[m] = df[m].fillna(0).astype(int)
            else:
                df[m] = 0
        sheets[name] = df

    sheets["Maintenance Charges"] = mc
    sheets["Payments"] = payments

    # Write with openpyxl so we can inject real formulas afterwards
    with pd.ExcelWriter(XLSX, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if name == "README":
                continue
            df.to_excel(writer, sheet_name=name, index=False)

    # Post-process: replace formula-placeholder strings with real Excel formulas
    from openpyxl import load_workbook

    wb = load_workbook(XLSX)
    mc_ws = wb["Maintenance Charges"]
    for row in range(2, mc_ws.max_row + 1):
        for month_idx in range(12):
            col_idx = 9 + month_idx  # I=9
            cell = mc_ws.cell(row=row, column=col_idx)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                # Replace {row} with actual Excel row number
                cell.value = cell.value.replace("{row}", str(row))

    wb.save(XLSX)
    print(f"Normalized {XLSX}")
    print(f"  Residents: {len(residents)}")
    print(f"  Maintenance Charges rows: {len(mc)}")
    print(f"  Payments sheet added")
    print(f"  Fine sheets cleaned and aligned")


if __name__ == "__main__":
    main()
