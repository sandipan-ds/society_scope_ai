"""Fix Maintenance Charges to hold the plain base amount (3500/month).

The workbook now has a separate 'Total Charges' sheet whose formula is
Maintenance + all fine sheets. Maintenance must therefore be the BASE only,
otherwise fines are double-counted (3500+fines+fines).

Also verifies the Total Charges sheet has SUM formulas in every month cell.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

XLSX = Path(r"C:\Users\sandi\Desktop\ML Working Folder\society_scope_ai\data\members_data\Housing_Society_Charges_and_Fines_Template_108_Residents.xlsx")

FINE_SHEETS = [
    "Parking Violation Fines",
    "Waste Management Fines",
    "Pet Policy Fines",
    "Noise Violation Fines",
    "Property Damage Fines",
    "Miscellaneous Fines",
]
BASE = 3500


def main() -> None:
    wb = load_workbook(XLSX)

    mc = wb["Maintenance Charges"]
    for row in range(2, mc.max_row + 1):
        for col in range(9, 21):  # I..T
            cell = mc.cell(row=row, column=col)
            if cell.value != BASE:
                cell.value = BASE
    print(f"Maintenance Charges: set {mc.max_row - 1} rows x 12 months to {BASE}")

    if "Total Charges" in wb.sheetnames:
        tc = wb["Total Charges"]
        fixed = 0
        for row in range(2, tc.max_row + 1):
            for col in range(9, 21):
                cell = tc.cell(row=row, column=col)
                if not (isinstance(cell.value, str) and cell.value.startswith("=SUM")):
                    letter = cell.column_letter
                    refs = ",\n".join(
                        [f"'Maintenance Charges'!{letter}{row}"]
                        + [f"'{name}'!{letter}{row}" for name in FINE_SHEETS]
                    )
                    cell.value = f"=SUM(\n{refs}\n)"
                    fixed += 1
        print(f"Total Charges: verified formulas ({fixed} cells filled)")

    wb.save(XLSX)
    print("Saved", XLSX)


if __name__ == "__main__":
    main()
